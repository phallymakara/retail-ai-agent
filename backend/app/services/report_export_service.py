import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def generate_excel_report(report_data: dict[str, Any]) -> bytes:
    """Generate a formatted Excel (.xlsx) workbook for inventory reports."""
    wb = Workbook()

    # Sheet 1: Summary & Overview
    ws_summary = wb.active
    ws_summary.title = "Inventory Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title styling
    title_font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="475569")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    bold_font = Font(name="Calibri", size=11, bold=True)
    border_thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    store_name = report_data.get("store_name", "All Store Branches")
    gen_time = report_data.get("generated_at", datetime.now().isoformat())

    ws_summary.append(["RETAIL INVENTORY SUMMARY REPORT"])
    ws_summary.append([f"Store Branch: {store_name} | Generated: {gen_time}"])
    ws_summary.append([])

    ws_summary.cell(row=1, column=1).font = title_font
    ws_summary.cell(row=2, column=1).font = subtitle_font

    # Summary Metrics Table
    ws_summary.append(["Key Inventory Metric", "Value"])
    summary_row_start = ws_summary.max_row
    for col in range(1, 3):
        cell = ws_summary.cell(row=summary_row_start, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws_summary.append(["Total Products Tracked", report_data.get("total_products_tracked", 0)])
    ws_summary.append(["Total Stock Quantity", report_data.get("total_stock_quantity", 0)])
    ws_summary.append(["Total Available Quantity", report_data.get("total_available_quantity", 0)])
    ws_summary.append(["Total Reserved Quantity", report_data.get("total_reserved_quantity", 0)])
    ws_summary.append(["Low Stock Items Count", report_data.get("low_stock_count", 0)])
    ws_summary.append(["Out of Stock Items Count", report_data.get("out_of_stock_count", 0)])
    ws_summary.append([])

    # Apply borders & font styling to metrics
    for r in range(summary_row_start, ws_summary.max_row):
        for c in range(1, 3):
            cell = ws_summary.cell(row=r, column=c)
            cell.border = border_thin
            if c == 1 and r > summary_row_start:
                cell.font = bold_font

    # Sheet 2: Product Breakdown
    ws_breakdown = wb.create_sheet(title="Product Breakdown")
    ws_breakdown.views.sheetView[0].showGridLines = True

    ws_breakdown.append(["Product Breakdown by Category & Items"])
    ws_breakdown.cell(row=1, column=1).font = title_font
    ws_breakdown.append([])

    headers = ["Category", "Product SKU", "Product Name", "Available Stock Quantity"]
    ws_breakdown.append(headers)
    bd_row_start = ws_breakdown.max_row
    for col in range(1, len(headers) + 1):
        cell = ws_breakdown.cell(row=bd_row_start, column=col)
        cell.font = header_font
        cell.fill = header_fill

    category_breakdown = report_data.get("category_breakdown", [])
    for cat_item in category_breakdown:
        cat_name = cat_item.get("category", "General")
        products = cat_item.get("products", [])
        if products:
            for p in products:
                ws_breakdown.append([
                    cat_name,
                    p.get("sku", ""),
                    p.get("product_name", ""),
                    p.get("available_quantity", 0),
                ])
        else:
            ws_breakdown.append([
                cat_name,
                "-",
                "Category Total",
                cat_item.get("available_quantity", 0),
            ])

    # Borders for breakdown
    for r in range(bd_row_start, ws_breakdown.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws_breakdown.cell(row=r, column=c).border = border_thin

    # Auto-fit column widths across sheets
    for sheet in [ws_summary, ws_breakdown]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_pdf_report(report_data: dict[str, Any]) -> bytes:
    """Generate a PDF document for inventory reports."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name="ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#1E3A8A"),
    )

    subtitle_style = ParagraphStyle(
        name="ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#64748B"),
    )

    section_style = ParagraphStyle(
        name="SectionHeader",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#1E40AF"),
        spaceBefore=12,
        spaceAfter=6,
    )

    body_style = ParagraphStyle(
        name="ReportBody",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#334155"),
    )

    elements = []

    # Title & Store Branch
    store_name = report_data.get("store_name", "All Store Branches")
    gen_time = report_data.get("generated_at", datetime.now().isoformat())

    elements.append(Paragraph("RETAIL INVENTORY SUMMARY REPORT", title_style))
    elements.append(Paragraph(f"<b>Branch:</b> {store_name} &nbsp;|&nbsp; <b>Generated:</b> {gen_time}", subtitle_style))
    elements.append(Spacer(1, 14))

    # Summary Tiles Table
    elements.append(Paragraph("SUMMARY OVERVIEW", section_style))

    metrics_data = [
        ["Total Tracked", "Available Stock", "Low Stock", "Out of Stock"],
        [
            str(report_data.get("total_products_tracked", 0)),
            str(report_data.get("total_available_quantity", 0)),
            str(report_data.get("low_stock_count", 0)),
            str(report_data.get("out_of_stock_count", 0)),
        ],
    ]

    metrics_table = Table(metrics_data, colWidths=[130, 130, 130, 130])
    metrics_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#475569")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, 1), 14),
            ("TEXTCOLOR", (0, 1), (0, 1), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor("#059669")),
            ("TEXTCOLOR", (2, 1), (2, 1), colors.HexColor("#D97706")),
            ("TEXTCOLOR", (3, 1), (3, 1), colors.HexColor("#DC2626")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
        ])
    )
    elements.append(metrics_table)
    elements.append(Spacer(1, 14))

    # Product Breakdown Table
    elements.append(Paragraph("PRODUCT BREAKDOWN BY CATEGORY", section_style))

    bd_table_data = [["Category", "SKU", "Product Name", "Available Stock"]]

    category_breakdown = report_data.get("category_breakdown", [])
    for cat_item in category_breakdown:
        cat_name = cat_item.get("category", "General")
        products = cat_item.get("products", [])
        if products:
            for idx, p in enumerate(products):
                c_label = cat_name if idx == 0 else ""
                bd_table_data.append([
                    c_label,
                    p.get("sku", ""),
                    p.get("product_name", ""),
                    str(p.get("available_quantity", 0)),
                ])
        else:
            bd_table_data.append([
                cat_name,
                "-",
                "Category Total",
                str(cat_item.get("available_quantity", 0)),
            ])

    if len(bd_table_data) > 1:
        bd_table = Table(bd_table_data, colWidths=[120, 100, 215, 85])
        bd_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (2, -1), "LEFT"),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ])
        )
        elements.append(bd_table)

    elements.append(Spacer(1, 16))

    # Action Recommendations
    elements.append(Paragraph("RECOMMENDED NEXT ACTIONS FOR STORE STAFF", section_style))
    low_cnt = report_data.get("low_stock_count", 0)
    out_cnt = report_data.get("out_of_stock_count", 0)

    rec_texts = []
    if out_cnt > 0:
        rec_texts.append(f"• <b>Priority Restock Needed:</b> {out_cnt} product(s) currently out of stock. Initiate stock adjustments or inter-branch transfer proposals immediately.")
    if low_cnt > 0:
        rec_texts.append(f"• <b>Reorder Threshold Warning:</b> {low_cnt} product(s) approaching or below reorder limits. Review supplier reorders.")
    if out_cnt == 0 and low_cnt == 0:
        rec_texts.append("• <b>Stock Levels Healthy:</b> All tracked inventory items are sufficiently stocked across branches.")

    for r_text in rec_texts:
        elements.append(Paragraph(r_text, body_style))
        elements.append(Spacer(1, 4))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
